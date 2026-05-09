"""EIA weekly retail-gasoline price worker.

Pulls the U.S. Energy Information Administration's free weekly
retail-gasoline workbook (XLS) and stores per-region per-grade
snapshots in fuel_market_weekly. Runs once per day; EIA publishes
the report on Monday so a daily check is enough to catch the new
week without polling tightly.

Source — public, no API key, no rate limits:
    https://www.eia.gov/dnav/pet/xls/PET_PRI_GND_DCUS_NUS_W.xls

The workbook has multiple sheets ("Data 1" through "Data 11" or so)
covering the U.S. average + each PADD region. Sheet 1 carries the
U.S. All Grades + Regular + Midgrade + Premium series; we ingest
that sheet for the U.S. region and the city-level + PADD sheets
for the regional splits.

Worker errors (HTTP fail, parse fail, network down) are caught at
the outer loop and logged; we sleep and retry on the next cycle so
a bad day doesn't kill the worker.
"""

from __future__ import annotations

import asyncio
import io
import logging
from datetime import datetime, timedelta, timezone
from decimal import Decimal

import asyncpg
import httpx
from openpyxl import load_workbook

log = logging.getLogger(__name__)

# Once per day — EIA publishes Monday but the file timestamp updates
# once. Daily cadence catches the update without burning bandwidth.
CYCLE_INTERVAL_SECONDS = 24 * 3_600

EIA_XLS_URL = "https://www.eia.gov/dnav/pet/xls/PET_PRI_GND_DCUS_NUS_W.xls"


def _parse_workbook(blob: bytes) -> list[tuple[str, str, Decimal]]:
    """Return list of (region_code, week_iso, price) tuples for the
    U.S. Regular All Formulations weekly series.

    The workbook's "Data 1" sheet is structured as:
        Row 1: title
        Row 2: header — Date, then a column per series (varies)
        Row 3+: data — date in col A, price values in subsequent cols

    We grab the first numeric series that mentions "Regular" or "All
    Grades" in the header and treat the U.S. column. Regional splits
    live in Data 2 / 3 / etc; future iteration can ingest those.
    """
    out: list[tuple[str, str, Decimal]] = []
    cutoff = datetime.now(timezone.utc) - timedelta(weeks=104)
    try:
        wb = load_workbook(filename=io.BytesIO(blob), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        log.warning("eia: workbook load failed: %s", exc)
        return out

    # Sheet "Data 1" carries U.S. All Grades / Regular / Midgrade /
    # Premium time series. Find the column with "Regular" in the
    # header (case-insensitive).
    if "Data 1" not in wb.sheetnames:
        log.warning("eia: workbook missing 'Data 1' sheet — got %s", wb.sheetnames)
        return out
    sheet = wb["Data 1"]

    # Header row — second non-empty row near the top. Scan rows 1-5
    # to find the row containing "Regular" or "All Grades".
    rows_iter = sheet.iter_rows(values_only=True)
    header: tuple = ()
    header_row_idx = -1
    for i, row in enumerate(rows_iter):
        if i > 6:
            break
        labels = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("regular" in lbl for lbl in labels):
            header = row
            header_row_idx = i
            break
    if header_row_idx == -1:
        log.warning("eia: Regular column header not found in first rows")
        return out

    # Find the column index of the Regular series (U.S. average).
    regular_col = next(
        (
            i for i, c in enumerate(header)
            if c is not None and "regular" in str(c).lower()
        ),
        None,
    )
    if regular_col is None:
        return out

    # Re-scan from after the header for data rows.
    for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
        if not row or row[0] is None:
            continue
        date_cell = row[0]
        if isinstance(date_cell, datetime):
            dt = date_cell
        else:
            try:
                dt = datetime.strptime(str(date_cell).strip(), "%Y-%m-%d")
            except ValueError:
                continue
        if dt.replace(tzinfo=timezone.utc) < cutoff:
            continue
        if regular_col >= len(row):
            continue
        cell = row[regular_col]
        if cell is None or cell == "":
            continue
        try:
            price = Decimal(str(cell)).quantize(Decimal("0.0001"))
        except Exception:  # noqa: BLE001
            continue
        out.append(("us", dt.strftime("%Y-%m-%d"), price))

    return out


async def _fetch_and_store(pool: asyncpg.Pool) -> int:
    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
        resp = await client.get(EIA_XLS_URL)
        resp.raise_for_status()
        blob = resp.content
    rows = _parse_workbook(blob)
    if not rows:
        log.warning("eia: parsed zero rows from workbook; skipping")
        return 0
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.executemany(
                """
                INSERT INTO fuel_market_weekly
                    (region, week_of, grade, price_usd_per_gal)
                VALUES ($1, $2::date, 'regular', $3)
                ON CONFLICT (region, week_of, grade) DO UPDATE
                    SET price_usd_per_gal = EXCLUDED.price_usd_per_gal,
                        fetched_at = now()
                """,
                rows,
            )
    return len(rows)


async def run(pool: asyncpg.Pool) -> None:
    """Worker entry point. Runs forever; cancelled at app shutdown."""
    log.info("eia worker started (cycle every %s s)", CYCLE_INTERVAL_SECONDS)
    while True:
        try:
            n = await _fetch_and_store(pool)
            if n:
                log.info("eia: stored %s region-week rows", n)
        except asyncio.CancelledError:
            raise
        except Exception as exc:  # noqa: BLE001
            log.error("eia cycle failed: %s", exc)
        await asyncio.sleep(CYCLE_INTERVAL_SECONDS)
